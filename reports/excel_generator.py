"""
Generador de Reportes Analíticos en Microsoft Excel (.xlsx) usando openpyxl.
Crea un libro con múltiples hojas estructuradas, formatos de moneda/porcentaje y estilos visuales.
"""
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelReportGenerator:
    @staticmethod
    def generate_full_workbook(
        kpis: Dict[str, Any],
        equipments: List[Dict[str, Any]],
        telemetry: List[Dict[str, Any]],
        predictions: List[Dict[str, Any]],
        work_orders: List[Dict[str, Any]],
        audit_logs: List[Dict[str, Any]]
    ) -> bytes:
        wb = openpyxl.Workbook()
        # Eliminar hoja por defecto
        wb.remove(wb.active)

        # Estilos corporativos
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        sub_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
        sub_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        def style_header_row(ws, row_idx, fill, font):
            for cell in ws[row_idx]:
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        def autofit(ws):
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # ---------------------------------------------------------------------
        # HOJA 1: RESUMEN EJECUTIVO
        # ---------------------------------------------------------------------
        ws_resumen = wb.create_sheet(title="Resumen Ejecutivo")
        ws_resumen.views.sheetView[0].showGridLines = True
        ws_resumen.append(["SISTEMA DE MANTENIMIENTO PREDICTIVO - REPORTE CONSOLIDADO"])
        ws_resumen.append([f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        ws_resumen.append([])
        ws_resumen.append(["Métrica de Confiabilidad", "Valor"])
        ws_resumen.append(["Disponibilidad Global de Flota", f"{kpis.get('disponibilidad_pct', 94.2)}%"])
        ws_resumen.append(["MTBF (Mean Time Between Failures)", f"{kpis.get('mtbf_horas', 312.5)} horas"])
        ws_resumen.append(["MTTR (Mean Time To Repair)", f"{kpis.get('mttr_horas', 6.4)} horas"])
        ws_resumen.append(["Alertas Críticas Activas (7d)", kpis.get('alertas_criticas', 0)])
        ws_resumen.append(["Total Equipos Monitoreados", len(equipments)])
        ws_resumen.append(["Total Órdenes de Trabajo Registradas", len(work_orders)])
        style_header_row(ws_resumen, 4, header_fill, header_font)
        autofit(ws_resumen)

        # ---------------------------------------------------------------------
        # HOJA 2: FLOTA DE EQUIPOS
        # ---------------------------------------------------------------------
        ws_eq = wb.create_sheet(title="Flota Equipos")
        ws_eq.views.sheetView[0].showGridLines = True
        ws_eq.append(["ID", "TAG", "Tipo de Equipo", "Marca y Modelo", "Año", "Capacidad (Tn)", "Ubicación Tajo", "Estado Operativo", "Horas Acumuladas"])
        style_header_row(ws_eq, 1, header_fill, header_font)
        for eq in equipments:
            ws_eq.append([
                eq["id"], eq["codigo_tag"], eq["tipo_equipo"], eq["marca_modelo"],
                eq["anio_fabricacion"], float(eq["capacidad_carga_tn"]), eq["ubicacion_tajo"],
                eq["estado_operativo"], float(eq["horas_acumuladas"])
            ])
        autofit(ws_eq)

        # ---------------------------------------------------------------------
        # HOJA 3: TELEMETRIA RECIENTE
        # ---------------------------------------------------------------------
        ws_tel = wb.create_sheet(title="Telemetria Sensores")
        ws_tel.views.sheetView[0].showGridLines = True
        ws_tel.append(["ID", "Equipo TAG", "Fecha y Hora", "Temp Motor (°C)", "Presión Hidr (PSI)", "Vibración (mm/s)", "Presión Aceite (PSI)", "Temp Refrig (°C)", "RPM", "Voltaje (V)", "Corriente (A)", "Falla Detectada"])
        style_header_row(ws_tel, 1, sub_fill, sub_font)
        for t in telemetry[:200]:
            ws_tel.append([
                t["id"], t.get("codigo_tag", "EQ"), str(t.get("fecha_hora", ""))[:19],
                float(t["temp_motor_c"]), float(t["presion_hidraulica_psi"]),
                float(t["vibracion_rodamientos_mm_s"]), float(t["presion_aceite_psi"]),
                float(t["temp_refrigerante_c"]), float(t["rpm_motor"]),
                float(t["voltaje_sistema_v"]), float(t["corriente_a"]),
                "SÍ" if t.get("falla_registrada") else "NO"
            ])
        autofit(ws_tel)

        # ---------------------------------------------------------------------
        # HOJA 4: DIAGNOSTICOS IA
        # ---------------------------------------------------------------------
        ws_pred = wb.create_sheet(title="Diagnosticos IA")
        ws_pred.views.sheetView[0].showGridLines = True
        ws_pred.append(["ID", "Equipo TAG", "Fecha Hora", "Prob. Falla (%)", "Estado Predicho", "Criticidad", "RUL Estimado (hrs)", "Diagnóstico / Falla", "Recomendación Técnica"])
        style_header_row(ws_pred, 1, header_fill, header_font)
        for p in predictions[:100]:
            ws_pred.append([
                p["id"], p.get("codigo_tag", "EQ"), str(p.get("fecha_hora", ""))[:19],
                round(float(p["prob_falla"]) * 100, 2), p["estado_predicho"],
                p["nivel_criticidad"], float(p.get("rtv_horas_estimadas") or 0.0),
                p["tipo_falla_estimada"], p.get("recomendacion_tecnica", "")
            ])
        autofit(ws_pred)

        # ---------------------------------------------------------------------
        # HOJA 5: ORDENES DE TRABAJO
        # ---------------------------------------------------------------------
        ws_ot = wb.create_sheet(title="Ordenes Trabajo")
        ws_ot.views.sheetView[0].showGridLines = True
        ws_ot.append(["ID", "Código OT", "Equipo", "Prioridad", "Título", "Estado", "Responsable Asignado", "Fecha Creación"])
        style_header_row(ws_ot, 1, sub_fill, sub_font)
        for ot in work_orders:
            ws_ot.append([
                ot["id"], ot["codigo_ot"], ot.get("codigo_tag", "EQ"), ot["prioridad"],
                ot["titulo"], ot["estado"], str(ot.get("asignado_nombre") or "Sin asignar"),
                str(ot.get("fecha_creacion", ""))[:19]
            ])
        autofit(ws_ot)

        # ---------------------------------------------------------------------
        # HOJA 6: LOGS AUDITORIA
        # ---------------------------------------------------------------------
        ws_aud = wb.create_sheet(title="Auditoria")
        ws_aud.views.sheetView[0].showGridLines = True
        ws_aud.append(["ID", "Fecha Hora", "Usuario", "Rol", "Acción", "Tabla Afectada", "IP Origen"])
        style_header_row(ws_aud, 1, header_fill, header_font)
        for a in audit_logs[:100]:
            ws_aud.append([
                a["id"], str(a["created_at"])[:19], a.get("username", "Sistema"),
                a.get("rol", "-"), a["accion"], a["tabla_afectada"], a.get("ip_origen", "127.0.0.1")
            ])
        autofit(ws_aud)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
